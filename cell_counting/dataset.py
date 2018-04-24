from concurrent import futures
import random, json, os, tempfile, shutil, importlib.util

from scipy import ndimage
import numpy as np
import tensorflow as tf
from scipy import misc
import openpyxl
import imgaug as ia

class Dataset(object):
    """A dataset consisting of some examples of input/output pairs. Minimizes
    memory usage by keeping most of the dataset on disk at any given time.

    Note that if there are extra examples beyond a multiple of SEGMENT_SIZE,
        then these examples will be ignored.
    """

    def __init__(self, path, segment_size=None):
        """Create a dataset.

        Note that only dataset sizes which are a multiple of segment_size will
            be allowed and any excess examples will be discarded.

        Args:
            path (str): The path to a directory in which to store this dataset,
                or if segment_size is omitted, the path to a directory in which
                a previously-stored dataset is stored.
            segment_size (int): The number of examples to store per file.  If
                omitted or None, this dataset will represent a previously-stored
                one (see description for path).
        """
        self._path = path
        if segment_size is None:
            self._load_metadata()
        else:
            self._segment_size = segment_size
            self._segments = 0
            self._save_metadata()

    def add(self, src):
        """Add the examples from a source dataset to this dataset.

        Args:
            src (Dataset): The source dataset.
        """

        def gen():
            for segment in range(src._segments):
                inputs, outputs = src._load_segment(segment)
                for i in range(inputs.shape[0]):
                    yield inputs[i, ...], outputs[i, ...]

        self._load_examples(gen())

    def load(self, loader_path):
        """Load examples using a loader.

        Args:
            loader_path (str): A path to a file that defines a function named
                ``load``. This function must yield the examples of the
                dataset.
        """
        spec = importlib.util.spec_from_file_location("loader", loader_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        self._load_examples(module.load())

    def initialize_from_aspects(self, path, transform):
        """Initialize by reading some examples' aspects and generating inputs
            and outputs from them.

            Each directory in the data directory will be interpreted as
                representing an example. In each of these example directories,
                each file will be interpreted as representing an aspect of the
                example, with the aspect's name being the file's base name and
                the aspect's value being the file's contents. Each aspect file
                must be in either a PNG or CSV format, with the appropriate file
                extension. Images must be 8-bit grayscale.

        Args:
            path (str): The path to the data directory.
            transform (func): A function that when called with a dictionary
                mapping aspect names to aspect values for a particular example
                returns a tuple of input np.ndarray and output np.ndarray for
                that example.
        """
        def get_example(dir_ent):
            if not dir_ent.is_dir():
                return
            aspects = {}
            for example_dir_ent in os.scandir(dir_ent.path):
                if example_dir_ent.is_dir():
                    continue
                name, _ = os.path.splitext(example_dir_ent.name)
                name = name.lower()
                val = self._read_aspect(example_dir_ent.path)
                aspects[name] = val
            return transform(aspects)
        inputs, outputs = [], []
        with futures.ThreadPoolExecutor() as executor:
            for example in executor.map(get_example, os.scandir(path)):
                if example is None:
                    continue
                inp, out = example
                inputs.append(inp)
                outputs.append(out)
        def example_iterable():
            for inp, out in zip(inputs, outputs):
                yield inp, out
        self._load_examples(example_iterable())

    _aspect_readers = {}

    @staticmethod
    def _read_aspect(path):
        _, ext = os.path.splitext(path)
        ext = ext.lower()
        if ext in Dataset._aspect_readers:
            return Dataset._aspect_readers[ext](path)
        else:
            raise ValueError("aspect file at {0:s} is not of supported type" \
                .format(path))

    def _read_aspect_png(path):
        return ndimage.imread(path, mode="L")

    _aspect_readers[".png"] = _read_aspect_png

    def _read_aspect_csv(path):
        return np.loadtxt(path, delimiter=',')

    _aspect_readers[".csv"] = _read_aspect_csv

    class ImageIntegerIterator(object):
        def __init__(self, image_dir_path, labels, shape):
            self._image_dir_path = image_dir_path
            self._labels = labels
            self._shape = shape
            self._image_name_iterator = iter(os.listdir(self._image_dir_path))

        def __iter__(self):
            return self

        def __next__(self):
            image_name = next(self._image_name_iterator)
            image_path = os.path.join(self._image_dir_path, image_name)
            image = ndimage.imread(image_path)
            image = misc.imresize(image, self._shape, interp="bilinear")
            image_name_noext = ".".join(image_name.split(".")[:-1])
            label = self._labels[image_name_noext]
            return (image, label)

    def load_images_and_excel_labels(self, image_dir_path, label_sheet_path,
            filename_col, label_col, shape):
        """Makes the inputs be the pixels of the images in some directory and
            the labels be given by an Excel spreadsheet mapping those images'
            filenames to integers.

        Args:
            image_dir_path (str): The path to the directory containing the
                images.
            label_sheet_path (str): The path to the spreadsheet containing the
                labels.
            filename_col (str): The column in the spreadsheet that contains
                the filenames.
            label_col (str): The column in the spreadsheet that contains the
                labels.
            shape (tuple of int): The shape to resize the images to. Should be
                a 3-element tuple of height, width, and channel depth.

        """
        labels = self._load_excel_mapping(label_sheet_path, filename_col,
            label_col)
        example_iterator = self.ImageIntegerIterator(image_dir_path, labels,
            shape)
        self._load_examples(example_iterator)

    class ImagesMaskPairIterator(object):
        def __init__(self, image_dir, mask_dir, dims):
            self._image_dir = image_dir
            self._mask_dir = mask_dir
            self._filenames = iter(os.listdir(self._mask_dir))
            self._dims = dims

        def __iter__(self):
            return self

        def __next__(self):
            filename = next(self._filenames)

            image_path = os.path.join(self._image_dir, filename)
            mask_path = os.path.join(self._mask_dir, filename)

            image = ndimage.imread(image_path)
            mask = ndimage.imread(mask_path)

            image = misc.imresize(image, self._dims, interp="bilinear")
            mask = misc.imresize(mask, self._dims, interp="bilinear")

            return image, mask

    def load_image_mask_pairs(self, image_dir, mask_dir, dims):
        """Makes the inputs be the pixels of the images in one directory and
            makes the outputs be the pixels of correspondingly-named mask images
            in another directory.

        Finds pairs of images and mask images by listing the files in the mask
            directory. Therefore, while all mask images need to have an image,
            not all images need to have a mask image.

        Args:
            image_dir (str): The path to the directory containing the images.
            mask_dir (str): The path to the directory containing the mask
                images.
            dims (tuple of int): The dimensions to resize the images and mask
                images to. Should be a 2-element tuple of height and width.
        """
        examples = self.ImagesMaskPairIterator(image_dir, mask_dir, dims)
        self._load_examples(examples)

    class MaskedImagesFromMetadataIterator(object):
        MASK_EXCLUDE_MAX = 50

        def __init__(self, metadata, image_path_getter, mask_path_getter,
            label_getter, shape):
            self._metadata_iterator = iter(metadata.values())
            self._image_path_getter = image_path_getter
            self._mask_path_getter = mask_path_getter
            self._label_getter = label_getter
            self._shape = shape

        def __iter__(self):
            return self

        def __next__(self):
            example_metadata = next(self._metadata_iterator)

            image_path = self._image_path_getter(example_metadata)
            image = ndimage.imread(image_path)

            mask_path = self._mask_path_getter(example_metadata)
            mask = ndimage.imread(mask_path)

            mask = np.mean(mask, axis=2)
            image[mask < self.MASK_EXCLUDE_MAX, :] = 0
            image = misc.imresize(image, self._shape, interp="bilinear")

            label = self._label_getter(example_metadata)

            return (image, label)

    def load_images_masks_labels_from_json(self, metadata_path, image_path_getter,
        mask_path_getter, label_getter, shape):
        """Assumes there is a JSON metadata file which contains, for each
            example, the path of an image, the path of a mask image, and a
            label. Makes the inputs be the images with pixels in the black area
            of the mask image set to black. Makes the outputs be the labels.

        Note that this will resize the mask images, which will lead to mask
            image values that are neither black nor white. Pixels whose average
            value over R, G, and B is less than
            MaskedImagesFromMetadataIterator.MASK_EXCLUDE_MAX will be
            considered black.

        Args:
            metadata_path (str): The path to the JSON metadata file.
            image_path_getter (func(dict) -> str): A function that takes the
                metadata for an example and returns the path to the example's
                image.
            mask_path_getter (func(dict) -> str): A function that takes the JSON
                metadata for an example and returns the path to the example's
                mask image.
            label_getter (func(dict) -> int): A function that takes the JSON
                metadata for an example and returns the example's label.
            shape (tuple of int): The shape to resize the images and mask images
                to. Should be a 3-element tuple of height, width, and channel
                depth.
        """
        metadata = json.load(open(metadata_path))
        example_iterator = self.MaskedImagesFromMetadataIterator(metadata,
            image_path_getter, mask_path_getter, label_getter, shape)
        self._load_examples(example_iterator)

    def size(self):
        """Return the size of this dataset.

        Returns:
            (int) The size of this dataset.
        """
        return self._segments * self._segment_size

    def split(self, p, path_larger, path_smaller, seed=None):
        """Split this dataset.

        The larger dataset will be given a segment size that is ``1 - p`` times
        this dataset's segment size, and likewise with ``p`` for the smaller
        dataset. Chooses the examples to allocate to each side of the split
        at random.

        Args:
            p (float): The proportion of the examples to put in the smaller part
                of the split.
            path_larger (str): The path to a directory in which to store the
                larger part of the split.
            path_smaller (str): The path to a directory in which to store the
                smaller part of the split.
            seed (int): A number to seed the random number generator with before
                randomly assigning examples to splits. With the same ``seed``,
                two ``split`` operations should have the same effect.

        Returns:
            tuple(dataset.Dataset, dataset.Dataset): The two datasets that
            result from the split. The smaller one is last.
        """
        def new_segment_size(p):
            return max(1, round(self._segment_size * p))

        larger = Dataset(path_larger, new_segment_size(1 - p))
        smaller = Dataset(path_smaller, new_segment_size(p))

        def example_generator(include_example):
            i = 0
            for segment in range(self._segments):
                inputs_path = self._get_segment_file_path(segment, "inputs.npy")
                outputs_path = self._get_segment_file_path(segment,
                                                           "outputs.npy")
                inputs = np.load(inputs_path)
                outputs = np.load(outputs_path)
                for segment_i in range(inputs.shape[0]):
                    if include_example(i):
                        yield (inputs[segment_i, ...], outputs[segment_i, ...])
                    i += 1

        num_examples = self._segments * self._segment_size
        num_examples_smaller = round(num_examples * p)
        if seed is not None:
            random.seed(seed)
        chosen_for_smaller = set(random.sample(range(num_examples),
                                               num_examples_smaller))

        def include_example_larger(i):
            return i not in chosen_for_smaller

        larger._load_examples(example_generator(include_example_larger))

        def include_example_smaller(i):
            return i in chosen_for_smaller

        smaller._load_examples(example_generator(include_example_smaller))
        return larger, smaller

    def map(self, fn):
        """Map a function onto this dataset.

        Args:
            fn (func(tuple of np.ndarray) -> sequence of tuple of np.ndarray):
                A function that takes in an example as a tuple of input and
                output. It returns a sequence of one or more new examples to
                replace the passed in example with.
        """
        def map_helper(batch):
            inputs, outputs = batch
            new_inputs = []
            new_outputs = []
            for i in range(inputs.shape[0]):
                example = (inputs[i, ...], outputs[i, ...])
                new_examples = fn(example)
                for new_example in new_examples:
                    new_inputs.append(new_example[0])
                    new_outputs.append(new_example[1])
            new_inputs = np.stack(new_inputs, axis=0)
            new_outputs = np.stack(new_outputs, axis=0)
            return (new_inputs, new_outputs)
        self.map_batch(map_helper)

    def map_generator(self, fn, save_path, segment_size=None):
        """Map a function onto the examples of this `Dataset`.

        Produces a new `Dataset`, leaving this one unmodified. This method
        tries to keep memory usage down by writing each complete segment of
        examples out to disk as soon it is produced.

        Args:
            fn (func): The map function. Takes as arguments the input and
                output associated with an example. Returns a generator over
                examples (``(input, output)`` pairs) that the example maps to.
            save_path (str): The save path for the new `Dataset`.
            segment_size (int): The segment size for the new `Dataset`. If
                omitted or ``None``, the segment size of this `Dataset` is used.

        Returns:
            (`Dataset`): A new `Dataset` containing the mapped examples.
        """
        if segment_size is None:
            segment_size = self._segment_size
        target = Dataset(save_path, segment_size)
        def example_generator():
            for segment_index in range(self._segments):
                inputs_path = self._get_segment_file_path(segment_index,
                                                          "inputs.npy")
                outputs_path = self._get_segment_file_path(segment_index,
                                                           "outputs.npy")
                inputs = np.load(inputs_path)
                outputs = np.load(outputs_path)
                for i in range(inputs.shape[0]):
                    inp, out = inputs[i, ...], outputs[i, ...]
                    yield from fn(inp, out)
        target._load_examples(example_generator())
        return target

    def map_batch(self, fn):
        """Map a function onto this dataset by applying it to batches.

        Args:
            fn (func(tuple of np.ndarray) -> tuple of np.ndarray):
                A function that takes in an array of inputs and an array of
                outputs for a batch of examples. It returns a new array of
                inputs and a new array of outputs to replace this batch, or
        """
        for segment_i in range(self._segments):
            inputs_path = self._get_segment_file_path(segment_i, "inputs.npy")
            outputs_path = self._get_segment_file_path(segment_i, "outputs.npy")
            inputs = np.load(inputs_path)
            outputs = np.load(outputs_path)
            new_inputs, new_outputs = fn((inputs, outputs))
            self._segment_size = new_inputs.shape[0]
            self._save_metadata()
            np.save(inputs_path, new_inputs)
            np.save(outputs_path, new_outputs)

    def set_segment_size(self, n):
        """Set the segment size of this dataset, the number of examples that
            this dataset stores in each of its files.

        Any extra examples beyond a multiple of the new segment size will be
            discarded.

        Args:
            n (int): The new segment size.
        """
        scratch_dataset = Dataset(tempfile.mkdtemp(), n)
        accum_inputs, accum_outputs = None, None
        next_segment = 0
        while next_segment < self._segments:
            # if not enough segments are accumulated, load one
            if accum_outputs is None or accum_outputs.shape[0] < n:
                seg_inputs, seg_outputs = self._load_segment(next_segment)
                if accum_inputs is None:
                    accum_inputs, accum_outputs = seg_inputs, seg_outputs
                else:
                    accum_inputs = np.concatenate((accum_inputs, seg_inputs),
                        axis=0)
                    accum_outputs = np.concatenate((accum_outputs, seg_outputs),
                        axis=0)
                next_segment += 1
            # if enough segments are accumulated, save one
            if accum_outputs.shape[0] >= n:
                # add only the first n if extra have been accumulated
                save_inputs = accum_inputs[:n, ...]
                accum_inputs = accum_inputs[n:, ...]
                save_outputs = accum_outputs[:n, ...]
                accum_outputs = accum_outputs[n:, ...]
                scratch_dataset._add_segment(save_inputs, save_outputs)
        # make this dataset be like the new dataset
        shutil.rmtree(self._path)
        shutil.copytree(scratch_dataset._path, self._path)
        scratch_dataset.close()
        self._segment_size = scratch_dataset._segment_size
        self._segments = scratch_dataset._segments
        self._save_metadata()

    def get_batch(self, size, pool_multiplier=5):
        """Get a batch of examples, randomly selected from a pool.

        Args:
            size (int): The number of examples.
            pool_multiplier (int): If ``n`` segments are needed to get ``size``
                examples, then the examples will be drawn at random from
                ``pool_multiplier * n`` segments. This result is then capped
                at the number of segments that exist. The default is ``5``.

        Returns:
            tuple(numpy.ndarray, numpy.ndarray): The inputs and outputs of
            the batch. Each array has ``size`` rows, where the ``i``-th row
            corresponds to the ``i``-th example.
        """
        assert isinstance(pool_multiplier, int), "argument pool_multiplier " \
                                                 "must be an int"
        assert pool_multiplier >= 1, "argument pool_multiplier must be at " \
                                     "least 1"

        segments_needed = (int(size / self._segment_size) + 1)
        segments_needed *= pool_multiplier
        segments_needed = min(segments_needed, self._segments)
        chosen_segments = random.sample(range(self._segments), segments_needed)
        cache = []
        for segment in chosen_segments:
            inputs_path = self._get_segment_file_path(segment, "inputs.npy")
            outputs_path = self._get_segment_file_path(segment, "outputs.npy")
            inputs = np.load(inputs_path)
            outputs = np.load(outputs_path)
            cache.append((inputs, outputs))
        chosen_examples = random.sample(
            range(segments_needed * self._segment_size), size)
        inputs = []
        outputs = []
        for example in chosen_examples:
            segment = int(example / self._segment_size)
            row = example % self._segment_size
            inputs.append(cache[segment][0][row, ...])
            outputs.append(cache[segment][1][row, ...])
        inputs = np.stack(inputs, axis=0)
        outputs = np.stack(outputs, axis=0)
        return inputs, outputs

    class BatchIterator:
        def __init__(self, dataset, batch_size, pool_size):
            self._dataset = dataset
            self._batch_size = batch_size
            self._pool_size = pool_size

        def __next__(self):
            return self._dataset.get_batch(self._batch_size, self._pool_size)

        def __iter__(self):
            return self

    class EpochBatchIterator:
        def __init__(self, dataset, batch_size, pool_size):
            self._dataset = dataset
            self._batch_size = batch_size
            self._pool_size = pool_size

            self._epoch = 0
            self._start_epoch()
            inputs, outputs = dataset.get_batch(2, 1)
            segments = (int(batch_size / dataset._segment_size) + 1)
            segments *= pool_size
            segments = min(segments, dataset._segments)
            self._pool_max = (segments + 2) * dataset._segment_size
            self._pool_min = segments * dataset._segment_size
            self._pool_inputs = np.empty((self._pool_max,) + inputs.shape[1:])
            self._pool_outputs = np.empty((self._pool_max,) + outputs.shape[1:])

        def __next__(self):
            self._replenish()
            chosen = np.random.choice(self._pool_top + 1, self._batch_size)
            inputs = self._pool_inputs[chosen, ...]
            outputs = self._pool_outputs[chosen, ...]
            new_inputs = np.delete(self._pool_inputs[:self._pool_top, ...],
                                   chosen, axis=0)
            new_outputs = np.delete(self._pool_outputs[:self._pool_top, ...],
                                    chosen, axis=0)
            self._pool_top = new_inputs.shape[0]
            self._pool_inputs[:self._pool_top, ...] = new_inputs
            self._pool_outputs[:self._pool_top, ...] = new_outputs
            return inputs, outputs

        def __iter__(self):
            return self

        def _replenish(self):
            while (self._pool_top + 1) <= self._pool_min:
                try:
                    segment = self._segment_use_map.index(False)
                except ValueError:
                    break
                self._segment_use_map[segment] = True
                inputs, outputs = self._dataset._load_segment(segment)
                new_pool_top = self._pool_top + inputs.shape[0]
                self._pool_inputs[self._pool_top:new_pool_top, ...] = inputs
                self._pool_outputs[self._pool_top:new_pool_top, ...] = outputs
                self._pool_top = new_pool_top
            if (self._pool_top + 1) < self._batch_size:
                self._start_epoch()
                self._replenish()

        def _start_epoch(self):
            self._epoch += 1
            self._pool_top = 0
            self._segment_use_map = [False for _
                                     in range(self._dataset._segments)]

        def get_epoch(self):
            return self._epoch

    def get_batch_iterable(self, batch_size, pool_size=5, epochs=False):
        """Get an iterable over batches of this dataset's examples.

        Args:
            batch_size (int): The number of examples per batch.
            pool_size (int): The minimum number of segments to pool together in
                memory to draw random batches from. If omitted, ``5``.
            epochs (bool): Whether to organize the batches into epochs, so that
                every example is seen once before any is  repeated,
                then every example is seen twice before any is repeated a third
                time, etc. If omitted, ``False``.

        Returns:
            (Iterable): An iterable over batches of this dataset's examples.
                Each element of the sequence is a ``tuple`` of
                ``numpy.ndarray``s, where the first array is the inputs and the
                second array is the outputs.
        """
        if not epochs:
            return Dataset.BatchIterator(self, batch_size, pool_size)
        else:
            return Dataset.EpochBatchIterator(self, batch_size, pool_size)

    def get_all(self):
        """Get all examples in the dataset.

        Returns the examples in the same order each time it is called.

        Returns:
            (tuple(numpy.ndarray, numpy.ndarray)): A tuple of two
                elements: the inputs of all the examples and the outputs of
                all the examples. These will have shape (num_examples, ...).
        """
        inputs, outputs = [], []
        for i in range(self._segments):
            inp, out = self._load_segment(i)
            inputs.append(inp)
            outputs.append(out)
        inputs = np.concatenate(inputs, axis=0)
        outputs = np.concatenate(outputs, axis=0)
        assert inputs.shape[0] == outputs.shape[0], "Something is wrong!" \
            "There are not equal numbers of inputs and outputs in this dataset."
        return (inputs, outputs)

    def get_data_fn(self, batch_size, num_batches):
        """Get a data function for this dataset.

        Args:
            batch_size (int): The number of examples to put in each batch.
            num_batches (int): The number of batches to produce.

        Returns:
            (func): A data function giving num_batches batches of batch_size
                examples each.
        """
        inputs, outputs = self.get_batch(batch_size * num_batches)
        assert not np.any(np.isnan(inputs))
        assert not np.any(np.isnan(outputs))
        return tf.estimator.inputs.numpy_input_fn({"inputs":inputs}, outputs,
            batch_size, num_batches, shuffle=False,
            queue_capacity=num_batches)

    def delete(self):
        """Delete this dataset.

        Deletes this dataset's directory on disk. This will cause any files
        in that directory, even those not created by this dataset,
        to be deleted. This dataset will not be usable afterward."""
        shutil.rmtree(self._path)
        self._path = None

    def augment(self, augmenter=None, input_augmenter=None,
                output_augmenter=None):
        """Augment this dataset.

        Applies transformations to the examples in this dataset to produce
            additional examples.

        Args:
            augmenter (imgaug.augmenters.Augmenter): An augmenter.
            augment_inputs (bool): Whether to apply the augmenter to the inputs.
            augment_outputs (bool): Whether to apply the augmenter to the
                outputs.
        """
        original_segments = self._segments
        self._segments = 0
        for i in range(original_segments):
            inputs, outputs = self._load_segment(i)
            if augmenter is not None:
                det_augmenter = augmenter.to_deterministic()
                gen = det_augmenter.augment_batches([inputs])
                inputs = list(gen)[0]
                gen = det_augmenter.augment_batches([outputs])
                outputs = list(gen)[0]
            if input_augmenter is not None:
                gen = input_augmenter.augment_batches([inputs])
                inputs = list(gen)[0]
            if output_augmenter is not None:
                gen = output_augmenter.augment_batches([outputs])
                outputs = list(gen)[0]
            shutil.rmtree(self._get_segment_path(self._segments))
            self._add_segment(inputs, outputs)

    def _load_examples(self, example_iterator):
        inputs, outputs = [], []
        for i, example in enumerate(example_iterator):
            input, output = example
            inputs.append(input)
            outputs.append(output)
            if (i + 1) % self._segment_size == 0:
                inputs = np.stack(inputs, axis=0)
                outputs = np.stack(outputs, axis=0)
                self._add_segment(inputs, outputs)
                inputs, outputs = [], []

    def _add_segment(self, inputs, outputs):
        segment_path = self._get_segment_path(self._segments)
        os.mkdir(segment_path)
        inputs_path = self._get_segment_file_path(self._segments, "inputs.npy")
        outputs_path = self._get_segment_file_path(self._segments,
            "outputs.npy")
        np.save(inputs_path, inputs)
        np.save(outputs_path, outputs)
        self._segments += 1
        self._save_metadata()

    def _load_segment(self, segment):
        inputs_path = self._get_segment_file_path(segment, "inputs.npy")
        outputs_path = self._get_segment_file_path(segment, "outputs.npy")
        inputs = np.load(inputs_path)
        outputs = np.load(outputs_path)
        return inputs, outputs

    def _get_segment_path(self, segment):
        return os.path.join(self._path, str(segment))

    def _get_segment_file_path(self, segment, filename):
        segment_path = self._get_segment_path(segment)
        segment_data_path = os.path.join(segment_path, filename)
        return segment_data_path

    def _save_metadata(self):
        meta_path = self._get_segment_path("meta")
        os.makedirs(meta_path, exist_ok=True)
        segment_size_path = self._get_segment_file_path("meta",
            "segment_size.npy")
        segments_path = self._get_segment_file_path("meta", "segments.npy")
        np.save(segment_size_path, self._segment_size)
        np.save(segments_path, self._segments)

    def _load_metadata(self):
        meta_path = self._get_segment_path("meta")
        segment_size_path = self._get_segment_file_path("meta",
            "segment_size.npy")
        segments_path = self._get_segment_file_path("meta", "segments.npy")
        self._segment_size = int(np.load(segment_size_path))
        self._segments = int(np.load(segments_path))

    @staticmethod
    def _load_excel_mapping(path, key_col, value_col):
        mapping = {}
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        i = 2
        while True:
            key = ws[key_col+str(i)].value
            val = ws[value_col+str(i)].value
            if key is None or val is None:
                return mapping
            mapping[key] = val
            i += 1