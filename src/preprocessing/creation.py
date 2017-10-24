import os, random, contextlib
from collections import Counter

from wand.image import Image
import openpyxl

def load_excel(path, name_column, data_column, limit=None):
	"""
	Load the labels written in some Excel file. Row 1 of the file must be a
		header row.

	Args:
		path (str): The path to the file.
		name_column (str): The column in which to find names. A single
			character.
		data_column (str): The column in which to find data. A single
			character.
		limit (int): The maximum number of labels to load. If not `None`, will
			load only the first `limit` labels.

	Returns:
		(dict of str:int): A dictionary from names to labels.
	"""
	if limit is None:
		limit = float("inf")
	labels = []
	wb = openpyxl.load_workbook(path)
	ws = wb.active
	i = 2
	while i < limit:
		filename = ws[name_column+str(i)].value
		label = ws[data_column+str(i)].value
		if filename is None or label is None:
			break
		labels.append((filename, label))
		i += 1
	return dict(labels)

def load_images(dir, limit=None):
	"""
	Load the images contained in some directory.

	Caution: This may leak resources! Proper usage of `Image` should actually
		wrap with a context manager!

	Args:
		dir (str): The path to the directory.
		limit (int): The maximum number of images to load. If not `None`, will
			load only the lexicographically first `limit` images.

	Returns:
		(dict of str:Image): A dictionary from filenames to images.
	"""
	paths = os.listdir(dir)
	if limit is not None:
		if len(paths) > limit:
			paths.sort()
			paths = paths[:limit]
	images = {}
	for path in paths:
		img = Image(filename=os.path.join(dir, path))
		images[path] = img
	return images

def standardize_dimensions(images, dims, filter="gaussian", blur=1):
	"""
	Make the dimensions of the images in a list all equal.

	Args:
		images (dict of str:Image): The images.
		dims (tuple of int): The dimensions to use.
		filter (str): The type of filter to use when rescaling. `gaussian` by
			default. See `wand.image.Image.resize`.
		blur (int): The amount of blurring to use when rescaling. `1` by
			default. See `wand.image.Image.resize`.
	"""
	for filename, img in images.items():
		img.resize(dims[0], dims[1], filter, blur)

def split(data, test_p=0.1, seed=42114):
	"""
	Split a set into a training set and a test set, randomly.

	By choosing the same seed for splitting the data and labels of a dataset,
		one can ensure that the data and labels get outputted in the same order
		and in the same subset.

	Args:
		data (dict of str:X): The data.
		test_p: (float): The proportion of the items which should go to the
			test set. `0.1` by default.
		seed (int): The seed to use for random number generation.

	Returns:
		(tuple of dict of str:X): Two sets of items, the training set and the
			test set.
	"""
	random.seed(seed)
	items = data.items()
	random.seed(seed)
	items.sort()
	random.shuffle(items)
	split_index = int(len(items) * test_p)
	train_items, test_items = items[split_index:], items[:split_index]
	return dict(train_items), dict(test_items)

def save_images(images, dir):
	"""
	Save a set of images to a directory using sequential numbers as filenames.
	
	Images are numbered in ascending order of their original filenames. Images
		are saved in their original format.

	Args:
		images (dict of str:Image): The images.
		dir (str): The path to the directory.
	"""
	items = images.items()
	items.sort()
	for i, item in enumerate(items):
		filename, img = item
		img.save(filename=os.path.join(dir, str(i) + "." + img.format))

def save_csv(data, path):
	"""
	Save some data to a CSV file using sequential numbers as identifiers.
	
	Items are numbered in ascending order of their filenames.

	Args:
		data (dict of str:X): The data.
		path (str): The path to save to.
	"""
	with open(path, "w+") as file:
		for i, item in enumerate(data.items()):
			_, label = item
			file.write(str(i))
			file.write(",")
			file.write(str(label))
			file.write("\n")
