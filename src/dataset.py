from scipy import ndimage
import numpy as np
import tempfile
import openpyxl
import os
from scipy import misc
import random
import json
import shutil

import tensorflow as tf
from sklearn.feature_extraction import image

class Dataset(object):
	"""A dataset consisting of some examples of input/output pairs. Minimizes
	memory usage by keeping most of the dataset on disk at any given time.

	Note that if there are extra examples beyond a multiple of SEGMENT_SIZE,
		then these examples will be ignored.
	"""

	SEGMENT_POOL_MULTIPLIER = 5

	def __init__(self, segment_size):
		"""Create a dataset.

		Args:
			segment_size (int): The number of examples to store per file. Note
				that only dataset sizes which are a multiple of segment_size
				will be allowed and any excess examples will be discarded.
		"""
		self._segment_size = segment_size
		self._segments = 0
		self._segment_dir = tempfile.TemporaryDirectory()

	class ImageIntegerIterator(object):
		def __init__(self, image_dir_path, labels, shape):
			self._image_dir_path = image_dir_path
			self._labels = labels
			self._shape = shape
			self._image_name_iterator = iter(os.listdir(self._image_dir_path))

		def __iter__(self):
			return self

		def __next__(self):
			image_name = next(self._image_name_iterator)
			image_path = os.path.join(self._image_dir_path, image_name)
			image = ndimage.imread(image_path)
			image = misc.imresize(image, self._shape, interp="bilinear")
			image_name_noext = ".".join(image_name.split(".")[:-1])
			label = self._labels[image_name_noext]
			return (image, label)

	def load_images_and_excel_labels(self, image_dir_path, label_sheet_path,
			filename_col, label_col, shape):
		"""Makes the inputs be the pixels of the images in some directory and
			the labels be given by an Excel spreadsheet mapping those images'
			filenames to integers.

		Args:
			image_dir_path (str): The path to the directory containing the
				images.
			label_sheet_path (str): The path to the spreadsheet containing the
				labels.
			filename_col (str): The column in the spreadsheet that contains
				the filenames.
			label_col (str): The column in the spreadsheet that contains the
				labels.
			shape (tuple of int): The shape to resize the images to. Should be
				a 3-element tuple of height, width, and channel depth.

		"""
		labels = self._load_excel_mapping(label_sheet_path, filename_col,
			label_col)
		example_iterator = self.ImageIntegerIterator(image_dir_path, labels,
			shape)
		self._load_examples(example_iterator)

	class ImagesMaskPairIterator(object):
		def __init__(self, image_dir, mask_dir, dims):
			self._image_dir = image_dir
			self._mask_dir = mask_dir
			self._filenames = iter(os.listdir(self._mask_dir))
			self._dims = dims

		def __iter__(self):
			return self

		def __next__(self):
			filename = next(self._filenames)

			image_path = os.path.join(self._image_dir, filename)
			mask_path = os.path.join(self._mask_dir, filename)

			image = ndimage.imread(image_path)
			mask = ndimage.imread(mask_path)

			image = misc.imresize(image, self._dims, interp="bilinear")
			mask = misc.imresize(mask, self._dims, interp="bilinear")

			return image, mask

	def load_image_mask_pairs(self, image_dir, mask_dir, dims):
		"""Makes the inputs be the pixels of the images in one directory and
			makes the outputs be the pixels of correspondingly-named mask images
			in another directory.

		Finds pairs of images and mask images by listing the files in the mask
			directory. Therefore, while all mask images need to have an image,
			not all images need to have a mask image.

		Args:
			image_dir (str): The path to the directory containing the images.
			mask_dir (str): The path to the directory containing the mask
				images.
			dims (tuple of int): The dimensions to resize the images and mask
				images to. Should be a 2-element tuple of height and width.
		"""
		examples = self.ImagesMaskPairIterator(image_dir, mask_dir, dims)
		self._load_examples(examples)

	class MaskedImagesFromMetadataIterator(object):
		MASK_EXCLUDE_MAX = 50

		def __init__(self, metadata, image_path_getter, mask_path_getter,
			label_getter, shape):
			self._metadata_iterator = iter(metadata.values())
			self._image_path_getter = image_path_getter
			self._mask_path_getter = mask_path_getter
			self._label_getter = label_getter
			self._shape = shape

		def __iter__(self):
			return self

		def __next__(self):
			example_metadata = next(self._metadata_iterator)

			image_path = self._image_path_getter(example_metadata)
			image = ndimage.imread(image_path)

			mask_path = self._mask_path_getter(example_metadata)
			mask = ndimage.imread(mask_path)

			mask = np.mean(mask, axis=2)
			image[mask < self.MASK_EXCLUDE_MAX, :] = 0
			image = misc.imresize(image, self._shape, interp="bilinear")

			label = self._label_getter(example_metadata)

			return (image, label)

	def load_images_masks_labels_from_json(self, metadata_path, image_path_getter,
		mask_path_getter, label_getter, shape):
		"""Assumes there is a JSON metadata file which contains, for each
			example, the path of an image, the path of a mask image, and a
			label. Makes the inputs be the images with pixels in the black area
			of the mask image set to black. Makes the outputs be the labels.

		Note that this will resize the mask images, which will lead to mask
			image values that are neither black nor white. Pixels whose average
			value over R, G, and B is less than
			MaskedImagesFromMetadataIterator.MASK_EXCLUDE_MAX will be
			considered black.

		Args:
			metadata_path (str): The path to the JSON metadata file.
			image_path_getter (func(dict) -> str): A function that takes the
				metadata for an example and returns the path to the example's
				image.
			mask_path_getter (func(dict) -> str): A function that takes the JSON
				metadata for an example and returns the path to the example's
				mask image.
			label_getter (func(dict) -> int): A function that takes the JSON
				metadata for an example and returns the example's label.
			shape (tuple of int): The shape to resize the images and mask images
				to. Should be a 3-element tuple of height, width, and channel
				depth.
		"""
		metadata = json.load(open(metadata_path))
		example_iterator = self.MaskedImagesFromMetadataIterator(metadata,
			image_path_getter, mask_path_getter, label_getter, shape)
		self._load_examples(example_iterator)

	def size(self):
		"""Return the size of this dataset.

		Returns:
			(int) The size of this dataset.
		"""
		return self._segments * self._segment_size

	def split(self, p):
		"""Split this dataset.

		Note: Will assign entire segments to one part of the split or the other.
			Result may not approximate p well if there aren't many segments.
			If the number of segments is very small, there is the possibility
			for errors.

		Args:
			p (float): The proportion of the dataset to put in the smaller part
				of the split.

		Returns:
			(tuple of dataset.Dataset): The two datasets that result from the
				split. The smaller one is last.
		"""
		smaller = Dataset(int(self._segment_size))
		larger = Dataset(int(self._segment_size))
		smaller_limit = int(p * self._segments)
		for i in range(smaller_limit):
			src = os.path.join(self._segment_dir.name, str(i))
			dst = os.path.join(smaller._segment_dir.name, str(i))
			shutil.copytree(src, dst)
		smaller._segments = smaller_limit
		for i in range(smaller_limit, self._segments):
			src = os.path.join(self._segment_dir.name, str(i))
			dst = os.path.join(larger._segment_dir.name, str(i - smaller_limit))
			shutil.copytree(src, dst)
		larger._segments = self._segments - smaller_limit
		return larger, smaller

	def map(self, fn):
		"""Map a function onto this dataset.

		Args:
			fn (func(tuple of np.ndarray) -> sequence of tuple of np.ndarray):
				A function that takes in an example as a tuple of input and
				output. It returns a sequence of one or more new examples to
				replace the passed in example with.
		"""
		def map_helper(batch):
			inputs, outputs = batch
			new_inputs = []
			new_outputs = []
			for i in range(inputs.shape[0]):
				example = (inputs[i, ...], outputs[i, ...])
				new_examples = fn(example)
				for new_example in new_examples:
					new_inputs.append(new_example[0])
					new_outputs.append(new_example[1])
			new_inputs = np.stack(new_inputs, axis=0)
			new_outputs = np.stack(new_outputs, axis=0)
			return (new_inputs, new_outputs)
		self.map_batch(map_helper)

	def map_batch(self, fn):
		"""Map a function onto this dataset by applying it to batches.

		Args:
			fn (func(tuple of np.ndarray) -> tuple of np.ndarray):
				A function that takes in an array of inputs and an array of
				outputs for a batch of examples. It returns a new array of
				inputs and a new array of outputs to replace this batch.
		"""
		for segment_i in range(self._segments):
			segment_dir = os.path.join(self._segment_dir.name, str(segment_i))
			inputs = np.load(os.path.join(segment_dir, "inputs.npy"))
			outputs = np.load(os.path.join(segment_dir, "outputs.npy"))
			new_inputs, new_outputs = fn((inputs, outputs))
			self._segment_size = new_inputs.shape[0]
			np.save(os.path.join(segment_dir, "inputs.npy"), new_inputs)
			np.save(os.path.join(segment_dir, "outputs.npy"), new_outputs)

	def set_segment_size(self, n):
		"""Set the segment size of this dataset, the number of examples that
			this dataset stores in each of its files.

		Any extra examples beyond a multiple of the new segment size will be
			discarded.

		Args:
			n (int): The new segment size.
		"""
		new_dataset = Dataset(n)
		accum_inputs, accum_outputs = None, None
		next_segment = 0
		while next_segment < self._segments:
			# if not enough segments are accumulated, load one
			if accum_outputs is None or accum_outputs.shape[0] < n:
				seg_inputs, seg_outputs = self._load_segment(next_segment)
				if accum_inputs is None:
					accum_inputs, accum_outputs = seg_inputs, seg_outputs
				else:
					accum_inputs = np.concatenate((accum_inputs, seg_inputs),
						axis=0)
					accum_outputs = np.concatenate((accum_outputs, seg_outputs),
						axis=0)
				next_segment += 1
			# if enough segments are accumulated, save one
			if accum_outputs.shape[0] >= n:
				# add only the first n if extra have been accumulated
				save_inputs = accum_inputs[:n, ...]
				accum_inputs = accum_inputs[n:, ...]
				save_outputs = accum_outputs[:n, ...]
				accum_outputs = accum_outputs[n:, ...]
				new_dataset._add_segment(save_inputs, save_outputs)
		# make this dataset be like the new dataset
		self.close()
		self._segment_size = new_dataset._segment_size
		self._segment_dir = new_dataset._segment_dir
		self._segments = new_dataset._segments

	def get_batch(self, size):
		"""Get a batch of examples.

		Args:
			size (int): The number of examples.

		Returns:
			(tuple of np.ndarray): The inputs and outputs of the batch. Each
				array has size rows, where the i-th row corresponds to the i-th
				example.
		"""
		segments_needed = (int(size / self._segment_size) + 1)
		segments_needed *= self.SEGMENT_POOL_MULTIPLIER
		segments_needed = min(segments_needed, self._segments)
		chosen_segments = random.sample(range(self._segments), segments_needed)
		cache = []
		for segment in chosen_segments:
			this_segment_dir = os.path.join(self._segment_dir.name,
				str(segment))
			inputs = np.load(os.path.join(this_segment_dir, "inputs.npy"))
			outputs = np.load(os.path.join(this_segment_dir, "outputs.npy"))
			cache.append((inputs, outputs))
		chosen_examples = random.sample(
			range(segments_needed * self._segment_size), size)
		inputs = []
		outputs = []
		for example in chosen_examples:
			segment = int(example / self._segment_size)
			row = example % self._segment_size
			inputs.append(cache[segment][0][row, ...])
			outputs.append(cache[segment][1][row, ...])
		inputs = np.stack(inputs, axis=0)
		outputs = np.stack(outputs, axis=0)
		return inputs, outputs

	def get_data_fn(self, batch_size, num_batches):
		"""Get a data function for this dataset.

		Args:
			batch_size (int): The number of examples to put in each batch.
			num_batches (int): The number of batches to produce.

		Returns:
			(func): A data function giving num_batches batches of batch_size
				examples each.
		"""
		inputs, outputs = self.get_batch(batch_size * num_batches)
		assert not np.any(np.isnan(inputs))
		assert not np.any(np.isnan(outputs))
		return tf.estimator.inputs.numpy_input_fn({"inputs":inputs}, outputs,
			batch_size, num_batches, shuffle=False,
			queue_capacity=num_batches)

	def close(self):
		"""Close this dataset. This dataset will not be useable afterward.

		This method must be called because it deletes temporary files on disk.
		"""
		self._segment_dir.__exit__(None, None, None)
		self._segment_dir = None

	def _load_examples(self, example_iterator):
		inputs, outputs = [], []
		for i, example in enumerate(example_iterator):
			input, output = example
			inputs.append(input)
			outputs.append(output)
			if (i + 1) % self._segment_size == 0:
				inputs = np.stack(inputs, axis=0)
				outputs = np.stack(outputs, axis=0)
				self._add_segment(inputs, outputs)
				inputs, outputs = [], []

	def _add_segment(self, inputs, outputs):
		this_segment_dir = os.path.join(self._segment_dir.name,
			str(self._segments))
		os.mkdir(this_segment_dir)
		np.save(os.path.join(this_segment_dir, "inputs.npy"), inputs)
		np.save(os.path.join(this_segment_dir, "outputs.npy"), outputs)
		self._segments += 1

	def _load_segment(self, segment):
		segment_dir = os.path.join(self._segment_dir.name, str(segment))
		inputs = np.load(os.path.join(segment_dir, "inputs.npy"))
		outputs = np.load(os.path.join(segment_dir, "outputs.npy"))
		return inputs, outputs

	@staticmethod
	def _load_excel_mapping(path, key_col, value_col):
		mapping = {}
		wb = openpyxl.load_workbook(path)
		ws = wb.active
		i = 2
		while True:
			key = ws[key_col+str(i)].value
			val = ws[value_col+str(i)].value
			if key is None or val is None:
				return mapping
			mapping[key] = val
			i += 1